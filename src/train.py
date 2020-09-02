import os
import matplotlib.pyplot as plt
import numpy as np
import torch
import torch.utils.data as data
import shutil

from dataset import Dataset
from model import get_classifier_model
from settings import TrainInformation, TrainResult
from utils import train_utils

plt.interactive(True)

prev_plot = 0


def set_optimizer(optimizer_method, model, init_lr, weight_decay, momentum=None):
    """Optimizer 설정."""
    if optimizer_method == "SGD":
        optimizer = torch.optim.SGD(
            model.parameters(),
            lr=init_lr,
            momentum=momentum,
            weight_decay=weight_decay,
            nesterov=True,
        )
    elif optimizer_method == "Adadelta":
        optimizer = torch.optim.Adadelta(
            model.parameters(), lr=init_lr, weight_decay=weight_decay
        )
    elif optimizer_method == "Adam":
        optimizer = torch.optim.Adam(
            model.parameters(), lr=init_lr, weight_decay=weight_decay
        )
    else:
        raise Exception("Unknown Optimizer {}".format(optimizer_method))
    return optimizer


def print_metrics(model, train_dataset, test_dataset, train_result):
    model.train(False)

    #test_preds = train_utils.get_preds(test_dataset.data[:, 1:], model)
    test_input = test_dataset.data[:, 1:]
    test_label = test_dataset.data[:, :1]
    test_preds = train_utils.get_preds(test_input, model)
    test_label_onehot = np.array([np.eye(int(np.max(test_label)+1), dtype=np.int_)[int(label)] for label in test_label])

    test_AUC = train_utils.compute_AUC(test_label_onehot, test_preds)
    test_AUC_per_class = train_utils.compute_AUC_per_class(test_label_onehot, test_preds)
    test_accuracy = train_utils.compute_accuracy(test_label, test_preds)


    #train_preds = train_utils.get_preds(train_dataset.data[:1000, 1:], model)
    train_input = train_dataset.data[:1000, 1:]
    train_label = train_dataset.data[:1000, :1]
    train_preds = train_utils.get_preds(train_input, model)
    train_label_onehot = np.array([np.eye(int(np.max(train_label)+1), dtype=np.int_)[int(label)] for label in train_label])

    train_AUC = train_utils.compute_AUC(train_label_onehot, train_preds)
    train_AUC_per_class = train_utils.compute_AUC_per_class(train_label_onehot, train_preds)
    train_accuracy = train_utils.compute_accuracy(train_label, train_preds)

    train_result.test_AUC_list.append("%.04f" % test_AUC)
    train_result.test_AUC_list_class.append(test_AUC_per_class)
    train_result.test_accuracy_list.append("%.04f" % test_accuracy)

    #print(f'train_AUC_per_class is {train_AUC_per_class}')
    #print(f'train_AUC is {train_AUC}')
    #print(f'test_AUC_per_class is {test_AUC_per_class}')
    #print(f'test_AUC is {test_AUC}')

    return train_AUC, test_AUC, train_accuracy, test_accuracy, test_preds


def compute_contributing_variables(model, test_dataset):
    print("Evaluating contributing variables")
    model.train(False)
    variable_by_column = np.load("../datasets/severity4_no_space_columnnames.npy")
    #variable_by_column = np.array([v.replace("HE_ast", "HE_alt") for v in variable_by_column])
    assert variable_by_column.shape[0] == test_dataset.data.shape[1] - 1, f'{variable_by_column.shape[0]}  {test_dataset.data.shape[1] - 1}'
    variables = np.unique(variable_by_column)
    AUCs = []
    print("Computing variable contributions")
    print(variables)
    for variable in variables:
        corresponding_indices = (variable_by_column == variable)
        #print("zeroing %s" % str(np.where(corresponding_indices)))
        val_data = test_dataset.data[:, 1:].copy()
        val_data[:, corresponding_indices] = 0.0
        #print((val_data[:, :17] ** 2).mean())
        #val_data = val_data * len(variables) / (len(variables) - 1)
        preds = train_utils.get_preds(val_data, model)
        target = test_dataset.data[:, :1]
        target_onehot = np.array([np.eye(int(np.max(target)+1), dtype=np.int_)[int(label)] for label in target])
        test_AUC = train_utils.compute_AUC(target_onehot, preds)
        print("%s %f" % (variable, test_AUC))
        AUCs.append(test_AUC)

    sorting_indices = np.argsort(AUCs)
    sorted_variables = [variables[i] for i in sorting_indices]
    sorted_AUCs = [AUCs[i] for i in sorting_indices]

    sorted_pairs = [(v, auc) for (v, auc) in zip(sorted_variables, sorted_AUCs)]
    for i, (v, auc) in enumerate(sorted_pairs[:20]):
        print("%03d: %s %f" % (i, v, auc))

    return [(v, auc) for (v, auc) in zip(variables, AUCs)]


def train_step(
        exp_name,
        ep,
        model,
        train_dataset,
        test_dataset,
        optimizer,
        init_lr,
        lr_decay,
        data_loader,
        bce_loss,
        train_result: TrainResult,
):
    global prev_plot
    model.train(True)
    for _, (X, y) in enumerate(data_loader):

        y = torch.tensor(y, dtype=torch.long)

        optimizer.zero_grad()
        pred_out = model(X.cuda()).view(X.shape[0], -1)
        #print(f'pred_out is {pred_out}')
        #print(f'pred_out.shape is {pred_out.shape}')
        loss = bce_loss(pred_out, y.cuda())
        loss.backward()
        avg_loss = train_result.avg_loss * 0.98 + loss.detach().cpu().numpy() * 0.02
        optimizer.step()
        train_result.total_iter += len(y)
        if train_result.total_iter % 10000 == 0:
            print(
                "Loss Iter %05d: %.4f\r" % (train_result.total_iter, avg_loss), end=""
            )
            train_result.loss_list.append(
                (train_result.total_iter, "{:.4f}".format(avg_loss))
            )
    print("")

    lr = init_lr * (lr_decay ** ep)
    for param_group in optimizer.param_groups:
        param_group["lr"] = lr
    print("Learning rate = %f" % lr)

    train_AUC, test_AUC, train_accuracy, test_accuracy, test_preds = print_metrics(model,
                                                                                   train_dataset,
                                                                                   test_dataset,
                                                                                   train_result)
    savedir = "/content/drive/My Drive/research/frontiers/checkpoints/%s" % exp_name
    os.makedirs(savedir, exist_ok=True)
    split = train_dataset.split
    savepath = "%s/epoch_%04d_fold_%02d.pt" % (savedir, ep, split)
    torch.save(model, savepath)

    if train_result.best_test_AUC < test_AUC:
        train_result.best_test_AUC = test_AUC
        train_result.best_test_epoch = ep
        if ep - prev_plot > 10:
            # 너무 자주 찍지 말고 한번 plot 찍고 epoch 10번 이상인 경우에만 찍는다.
            prev_plot = ep
            #train_utils.plot_AUC(test_dataset, test_preds, test_AUC)
        #contributing_variables = compute_contributing_variables(model, test_dataset)

    print(
        "Epoch %03d: test_AUC: %.4f (best: %.4f epoch: %d), train_AUC: %.4f"
        % (
            ep,
            test_AUC,
            train_result.best_test_AUC,
            train_result.best_test_epoch,
            train_AUC,
        )
    )
    print(
        "            test_accuracy {:.4f}, train_accuracy {:.4f}".format(
            test_accuracy, train_accuracy,
        )
    )


def train_logisticregressoin(info: TrainInformation, split, fold):
    """주어진 split에 대한 학습과 테스트를 진행한다."""
    bs = info.BS
    init_lr = info.INIT_LR
    lr_decay = info.LR_DECAY
    momentum = info.MOMENTUM
    weight_decay = info.WEIGHT_DECAY
    optimizer_method = info.OPTIMIZER_METHOD
    epoch = info.EPOCH
    nchs = info.NCHS
    filename = info.FILENAME
    model_name = info.MODEL_NAME
    exp_name = info.NAME

    print("Using File {}".format(filename))

    train_dataset = Dataset(split=split, fold=fold, phase="train", filename=filename, use_data_dropout=info.USE_DATA_DROPOUT)
    #val_dataset = Dataset(split=split, fold=fold, phase="val", filename=filename)
    test_dataset = Dataset(split=split, fold=fold, phase="test", filename=filename, use_data_dropout=False)

    import sklearn.linear_model

    regressor = sklearn.linear_model.LogisticRegression()
    regressor.fit(train_dataset.train_data[:, 1:], test_dataset.train_data[:, :1])
    preds = regressor.predict_proba(test_dataset.data[:, 1:])[:, 1]
    auc = train_utils.compute_AUC(test_dataset.data[:, :1], preds)
    print(auc)
    savepath = "/content/drive/My Drive/research/frontiers/checkpoints/logistic_regression/split_%02d.png" % split
    os.makedirs(os.path.dirname(savepath), exist_ok=True)
    #train_utils.plot_AUC_v2(preds, test_dataset.data[:, :1], savepath=savepath)

    model = get_classifier_model(model_name, train_dataset.feature_size, nchs, info.ACTIVATION)
    savedir = "/content/drive/My Drive/research/frontiers/checkpoints/%s" % exp_name
    best_test_epoch = 25
    loadpath = "%s/epoch_%04d_fold_%02d.pt" % (savedir, best_test_epoch, train_dataset.split)
    #model.load_state_dict(torch.load(savepath))
    model = torch.load(loadpath)
    model.eval()

    test_preds = train_utils.get_preds(test_dataset.data[:, 1:], model)
    train_utils.plot_AUC_v2([('Deep Neural Network', test_preds), ('Logistic Regression', preds)], test_dataset.data[:, :1], savepath=savepath)

def train_supportvectormachine(info: TrainInformation, split, fold):
    """주어진 split에 대한 학습과 테스트를 진행한다."""
    bs = info.BS
    init_lr = info.INIT_LR
    lr_decay = info.LR_DECAY
    momentum = info.MOMENTUM
    weight_decay = info.WEIGHT_DECAY
    optimizer_method = info.OPTIMIZER_METHOD
    epoch = info.EPOCH
    nchs = info.NCHS
    filename = info.FILENAME
    model_name = info.MODEL_NAME
    exp_name = info.NAME

    print("Using File {}".format(filename))

    train_dataset = Dataset(split=split, fold=fold, phase="train", filename=filename, use_data_dropout=info.USE_DATA_DROPOUT)
    #val_dataset = Dataset(split=split, fold=fold, phase="val", filename=filename)
    test_dataset = Dataset(split=split, fold=fold, phase="test", filename=filename, use_data_dropout=False)

    from sklearn.svm import LinearSVC

    regressor = LinearSVC()

    regressor.fit(train_dataset.train_data[:, 1:], test_dataset.train_data[:, :1])
    Y = regressor.decision_function(test_dataset.data[:, 1:])
    preds = (Y - Y.min()) / (Y.max() - Y.min())

    auc = train_utils.compute_AUC(test_dataset.data[:, :1], preds)
    print(auc)
    savepath = "/content/drive/My Drive/research/frontiers/checkpoints/logistic_regression/split_%02d.png" % split
    os.makedirs(os.path.dirname(savepath), exist_ok=True)
    #train_utils.plot_AUC_v2(preds, test_dataset.data[:, :1], savepath=savepath)

    model = get_classifier_model(model_name, train_dataset.feature_size, nchs, info.ACTIVATION)
    savedir = "/content/drive/My Drive/research/frontiers/checkpoints/%s" % exp_name
    best_test_epoch = 25
    loadpath = "%s/epoch_%04d_fold_%02d.pt" % (savedir, best_test_epoch, train_dataset.split)
    #model.load_state_dict(torch.load(savepath))
    model = torch.load(loadpath)
    model.eval()

    test_preds = train_utils.get_preds(test_dataset.data[:, 1:], model)
    train_utils.plot_AUC_v2([('Deep Neural Network', test_preds), ('support vector machine', preds)], test_dataset.data[:, :1], savepath=savepath)



def train_RandomForestClassifier(info: TrainInformation, split, fold):
    """주어진 split에 대한 학습과 테스트를 진행한다."""
    bs = info.BS
    init_lr = info.INIT_LR
    lr_decay = info.LR_DECAY
    momentum = info.MOMENTUM
    weight_decay = info.WEIGHT_DECAY
    optimizer_method = info.OPTIMIZER_METHOD
    epoch = info.EPOCH
    nchs = info.NCHS
    filename = info.FILENAME
    model_name = info.MODEL_NAME
    exp_name = info.NAME

    print("Using File {}".format(filename))

    train_dataset = Dataset(split=split, fold=fold, phase="train", filename=filename, use_data_dropout=info.USE_DATA_DROPOUT)
    #val_dataset = Dataset(split=split, fold=fold, phase="val", filename=filename)
    test_dataset = Dataset(split=split, fold=fold, phase="test", filename=filename, use_data_dropout=False)

    import sklearn.linear_model

    regressor = sklearn.linear_model.LogisticRegression()
    regressor.fit(train_dataset.train_data[:, 1:], test_dataset.train_data[:, :1])
    preds_regressor = regressor.predict_proba(test_dataset.data[:, 1:])[:, 1]
    auc_regressor = train_utils.compute_AUC(test_dataset.data[:, :1], preds_regressor)
    print(f'auc_regressor is {auc_regressor}')

    from sklearn.ensemble import RandomForestClassifier
    from sklearn.datasets import make_classification

    forest = RandomForestClassifier()
    forest.fit(train_dataset.train_data[:, 1:], test_dataset.train_data[:, :1])
    preds_forest = forest.predict_proba(test_dataset.data[:, 1:])[:, 1]
    auc_forest = train_utils.compute_AUC(test_dataset.data[:, :1], preds_forest)
    print(f'auc_forest is {auc_forest}')
    savepath = "/content/drive/My Drive/research/frontiers/checkpoints/random_forest/split_%02d.png" % split
    os.makedirs(os.path.dirname(savepath), exist_ok=True)
    #train_utils.plot_AUC_v2(preds, test_dataset.data[:, :1], savepath=savepath)

    model = get_classifier_model(model_name, train_dataset.feature_size, nchs, info.ACTIVATION)
    savedir = "/content/drive/My Drive/research/frontiers/checkpoints/%s" % exp_name
    best_test_epoch = 25
    loadpath = "%s/epoch_%04d_fold_%02d.pt" % (savedir, best_test_epoch, train_dataset.split)
    #model.load_state_dict(torch.load(savepath))
    model = torch.load(loadpath)
    model.eval()

    test_preds = train_utils.get_preds(test_dataset.data[:, 1:], model)
    train_utils.plot_AUC_v2([('Deep Neural Network', test_preds), ('Logistic Regression', preds_regressor), ('Random Forest', preds_forest)], test_dataset.data[:, :1], savepath=savepath)

def train_ml_compare(info: TrainInformation, split, fold):
    """주어진 split에 대한 학습과 테스트를 진행한다."""
    bs = info.BS
    init_lr = info.INIT_LR
    lr_decay = info.LR_DECAY
    momentum = info.MOMENTUM
    weight_decay = info.WEIGHT_DECAY
    optimizer_method = info.OPTIMIZER_METHOD
    epoch = info.EPOCH
    nchs = info.NCHS
    filename = info.FILENAME
    model_name = info.MODEL_NAME
    exp_name = info.NAME

    print("Using File {}".format(filename))

    train_dataset = Dataset(split=split, fold=fold, phase="train", filename=filename, use_data_dropout=info.USE_DATA_DROPOUT)
    #val_dataset = Dataset(split=split, fold=fold, phase="val", filename=filename)
    test_dataset = Dataset(split=split, fold=fold, phase="test", filename=filename, use_data_dropout=False)

    train_input = train_dataset.train_data[:, 1:]
    train_label = test_dataset.train_data[:, :1]

    # logisticregressoin ######################

    import sklearn.linear_model

    regressor = sklearn.linear_model.LogisticRegression()
    regressor.fit(train_input, train_label)
    preds_regressor = regressor.predict_proba(test_dataset.data[:, 1:])[:, 1]
    auc_regressor = train_utils.compute_AUC(test_dataset.data[:, :1], preds_regressor)
    print(f'auc_regressor is {auc_regressor}')

    ###########################################


    # randomforest ############################

    from sklearn.ensemble import RandomForestClassifier
    from sklearn.datasets import make_classification

    forest = RandomForestClassifier()
    forest.fit(train_input, train_label)
    preds_forest = forest.predict_proba(test_dataset.data[:, 1:])[:, 1]
    auc_forest = train_utils.compute_AUC(test_dataset.data[:, :1], preds_forest)
    print(f'auc_forest is {auc_forest}')

    ###########################################


    # svc #####################################

    from sklearn.svm import LinearSVC

    svc = LinearSVC()
    svc.fit(train_input, train_label)
    Y = svc.decision_function(test_dataset.data[:, 1:])
    preds_svc = (Y - Y.min()) / (Y.max() - Y.min())
    auc_svc = train_utils.compute_AUC(test_dataset.data[:, :1], preds_svc)
    print(f'auc_svc is {auc_svc}')

    ###########################################

    savepath = "/content/drive/My Drive/research/frontiers/checkpoints/ml_compare/split_%02d.png" % split
    os.makedirs(os.path.dirname(savepath), exist_ok=True)

    model = get_classifier_model(model_name, train_dataset.feature_size, nchs, info.ACTIVATION)
    savedir = "/content/drive/My Drive/research/frontiers/checkpoints/%s" % exp_name
    best_test_epoch = 25
    loadpath = "%s/epoch_%04d_fold_%02d.pt" % (savedir, best_test_epoch, train_dataset.split)
    #model.load_state_dict(torch.load(savepath))
    model = torch.load(loadpath)
    model.eval()

    test_preds = train_utils.get_preds(test_dataset.data[:, 1:], model)
    train_utils.plot_AUC_v2([('Deep Neural Network', test_preds), ('Logistic Regression', preds_regressor), ('Random Forest', preds_forest), ('Support Vector Classifier', preds_svc)], test_dataset.data[:, :1], savepath=savepath)

def train(info: TrainInformation, split, fold, combination, my_drive):
    """주어진 split에 대한 학습과 테스트를 진행한다."""
    """
    bs = info.BS
    init_lr = info.INIT_LR
    lr_decay = info.LR_DECAY
    momentum = info.MOMENTUM
    weight_decay = info.WEIGHT_DECAY
    optimizer_method = info.OPTIMIZER_METHOD
    nchs = info.NCHS
    model_name = info.MODEL_NAME
    epoch = info.EPOCH
    use_data_dropout = info.USE_DATA_DROPOUT
    activation = info.ACTIVATION
    """
    bs = combination[0]
    init_lr = combination[1]
    lr_decay = combination[2]
    momentum = combination[3]
    weight_decay = combination[4]
    optimizer_method = combination[5]
    nchs = combination[6]
    model_name = combination[7]
    epoch = combination[8]
    use_data_dropout = combination[9]
    activation = combination[10]

    exp_name = info.NAME

    filename = info.FILENAME

    print("Using File {}".format(filename))

    train_dataset = Dataset(split=split, fold=fold, phase="train", filename=filename, use_data_dropout=use_data_dropout)
    #val_dataset = Dataset(split=split, fold=fold, phase="val", filename=filename)
    test_dataset = Dataset(split=split, fold=fold, phase="test", filename=filename, use_data_dropout=False)

    test_input = test_dataset.data[:, 1:]
    test_label = test_dataset.data[:, :1]
    nchs[-1] = int(np.max(test_label)) + 1
    print(f'nchs is {nchs}')

    model = get_classifier_model(model_name, train_dataset.feature_size, nchs, activation)


    print(model)

    # Optimizer 설정
    optimizer = set_optimizer(
        optimizer_method, model, init_lr, weight_decay, momentum=momentum
    )

    data_loader = torch.utils.data.DataLoader(
        train_dataset, batch_size=bs, shuffle=True, num_workers=0, drop_last=True
    )

    #bce_loss = torch.nn.BCEWithLogitsLoss().cuda()
    bce_loss = torch.nn.CrossEntropyLoss().cuda()
    train_result = TrainResult()
    train_result.set_sizes(
        len(train_dataset.data), 0, len(test_dataset.data)
    )

    for ep in range(epoch):
        global prev_plot
        prev_plot = 0
        train_step(
            exp_name,
            ep,
            model,
            train_dataset,
            test_dataset,
            optimizer,
            init_lr,
            lr_decay,
            data_loader,
            bce_loss,
            train_result,
        )

    savedir = "/content/drive/My Drive/research/frontiers/checkpoints/%s" % exp_name
    best_test_epoch = train_result.best_test_epoch # 25
    savepath = "%s/epoch_%04d_fold_%02d.pt" % (savedir, best_test_epoch, train_dataset.split)
    #model.load_state_dict(torch.load(savepath))
    model = torch.load(savepath)
    model.eval()

    savepath_filenames = os.listdir(savdir)

    
    for savepath_filename in savepath_filenames:
        full_filename = os.path.join(savepath, savepath_filename)
        open(full_filename, 'w')


    shutil.rmtree(savedir)
    os.mkdir(savedir)
    # os.makedirs(savedir, exist_ok=True)

    """
    for a_file in my_drive.ListFile({'q': "trashed = true"}).GetList():
        # print the name of the file being deleted.
        print(f'the file {a_file["title"]}, is about to get deleted permanently.')
        # delete the file permanently.
        a_file.Delete()
    """

    test_input = test_dataset.data[:, 1:]
    test_label = test_dataset.data[:, :1]
    test_preds = train_utils.get_preds(test_input, model)
    test_label_onehot = np.array([np.eye(int(np.max(test_label)+1), dtype=np.int_)[int(label)] for label in test_label])
    test_AUC = train_utils.compute_AUC(test_label_onehot, test_preds)
    roc_auc = train_utils.plot_AUC_multi_class(test_dataset, test_preds, test_AUC, savepath=savepath.replace(".pt", "_AUC.png"))

    """
    contributing_variables = compute_contributing_variables(model, test_dataset)
    with open(os.path.join(savedir, "contributing_variables_epoch_%04d_fold_%02d.txt" % (best_test_epoch, train_dataset.split)), "w") as f:
        for (v, auc) in contributing_variables:
            f.write("%s %f\n" % (v, auc))
    """

    print(f'roc_auc is {roc_auc}')

    auc_class = []
    for class_num in range(int(np.max(test_label))+1):
        auc_class.append(roc_auc[class_num])

    print(f'auc_class is {auc_class}')

    info.split_index = split
    info.result_dict = train_result
    info.save_result()
    return train_result, auc_class


def run(filename, my_drive):

    """실험할 세팅을 불러오고, 그에 따라서 실험을 수행한다."""
    for a_file in my_drive.ListFile({'q': "trashed = true"}).GetList():
        # print the name of the file being deleted.
        print(f'the file {a_file["title"]}, is about to get deleted permanently.')
        # delete the file permanently.
        a_file.Delete()

    bs = [4096]  # 4096, 2048, 1024
    init_lr = [0.050, 0.025]  # 0.100000, 0.150000, 0.200000, 0.050000
    lr_decay = [0.999]  # 0.999, 0.99, 0.9, 0.8, 0.85
    momentum = [0.9]  # 0.9, 0.99, 0.999, 0.8, 0.85
    weight_decay = [1e-6]  # 1e-6, 1e-7, 5e-7, 5e-6
    optimizer_method = ["Adadelta"]  # "SGD", "Adadelta"
    nchs = [[2048, 2048, 2048, 2048, 2048, 2048, 512, 1], [2048, 2048, 2048, 2048, 2048, 512, 1]]  # [2048, 2048, 2048, 512, 1], [512, 512, 512, 512, 1], [512, 512, 1], [4096, 1]
    model_name = ["ClassifierWithAttention"]  # "ClassifierWithEmbedding", "Classifier", "ClassifierWithDropout", "ClassifierWithBatchNorm", "ClassifierWithAttention"
    epoch = [33]  # 26, 30, 40, 50
    use_data_dropout = ["use_data_dropout"]  # "use_data_dropout", None
    activation = ["swish"]  # "swish", "tanh", "ReLU", "LReLU"

    items = [bs, init_lr, lr_decay, momentum, weight_decay, optimizer_method, nchs, model_name, epoch, use_data_dropout, activation]

    from itertools import product
    combinations = list(product(*items))

    # cut combination number
    #combinations = combinations[100:]

    total_combination_number = len(combinations)
    print(f'We will check {total_combination_number} combinations')

    import openpyxl
    write_wb = openpyxl.Workbook()
    write_ws = write_wb.create_sheet('table')
    write_ws.append(['comb_index', 'bs', 'init_lr', 'lr_decay', 'momentum', 'weight_decay',
        'optimizer_method', 'nchs', 'model_name', 'epoch', 'use_data_dropout', 'activation',
        'best_test_auc', 'best_test_epoch', 'class 0', 'class 1', 'class 2', 'class 3'])
    write_wb.save('/content/drive/My Drive/research/frontiers/performance/performance.xlsx')

    for comb_index, combination in enumerate(combinations):
        info = TrainInformation(filename)
        np.random.seed(info.SEED)
        torch.manual_seed(info.SEED)
        fold = info.FOLD

        test_AUCs_by_split = []
        test_AUCs_by_split_class = []
        test_AUCs_by_split_class_test = []
        for split in range(fold):

            #if split % 3 > 0:
            #    print("Skipping split %d" % split)
            #    continue


            if False:
                #train_logisticregressoin(info, split, fold)
                #train_supportvectormachine(info, split, fold)
                train_ml_compare(info, split, fold)
                continue

            result, auc_class = train(info, split, fold, combination, my_drive)
            test_AUCs = [float(auc) for auc in result.test_AUC_list]
            test_AUCs_by_split.append(test_AUCs)
            test_AUCs_by_split_class.append(np.array(result.test_AUC_list_class))
            test_AUCs_by_split_class_test.append(auc_class)

            print(f'test_AUCs_by_split is {test_AUCs_by_split}')
            print(f'test_AUCs_by_split_class is {test_AUCs_by_split_class}')
            print(f'test_AUCs_by_split_class_test is {test_AUCs_by_split_class_test}')

            print(f'np.array(test_AUCs_by_split).shape is {np.array(test_AUCs_by_split).shape}')
            print(f'np.array(test_AUCs_by_split_class).shape is {np.array(test_AUCs_by_split_class).shape}')
            print(f'np.array(test_AUCs_by_split_class_test).shape is {np.array(test_AUCs_by_split_class_test).shape}')

        with open("result.txt", "a") as f:
            test_AUCs_by_split = np.array(test_AUCs_by_split)
            test_AUCs_by_epoch = test_AUCs_by_split.mean(axis=0)
            test_AUCs_by_split_class = np.array(test_AUCs_by_split_class)
            test_AUCs_by_epoch_class = np.transpose(test_AUCs_by_split_class.mean(axis=0))
            test_AUCs_by_split_class_test = np.array(test_AUCs_by_split_class_test)
            test_AUCs_by_epoch_class_test = test_AUCs_by_split_class_test.mean(axis=0)
            best_test_epoch = np.argmax(test_AUCs_by_epoch)
            best_test_AUC = test_AUCs_by_epoch[best_test_epoch]
            #f.write(str(info) + "/n")
            f.write("Name: %s\n" % info.NAME)
            f.write("average test AUC: %f %d\n" % (best_test_AUC, best_test_epoch))

            f.write("\n")
            f.write("best epoch\n")
            for class_number, mean_auc_per_class in enumerate(test_AUCs_by_epoch_class):
                f.write("average test AUC for %d class : %f %d\n" % (class_number, mean_auc_per_class[best_test_epoch], best_test_epoch))

            f.write("\n")
            f.write("test\n")
            for class_number, mean_auc_per_class in enumerate(test_AUCs_by_epoch_class_test):
                f.write("average test AUC for %d class : %f\n" % (class_number, mean_auc_per_class))

            f.write("\n")
            f.write("best\n")
            for class_number, mean_auc_per_class in enumerate(test_AUCs_by_epoch_class):
                best_test_epoch_class = np.argmax(mean_auc_per_class)
                f.write("average test AUC for %d class : %f %d\n" % (class_number, mean_auc_per_class[best_test_epoch_class], best_test_epoch_class))
            f.write("\n")
            f.write("\n")

        wb = openpyxl.load_workbook('/content/drive/My Drive/research/frontiers/performance/performance.xlsx')
        sheet1 = wb['table']
        print(f'combination is {combination}')
        print(f'best_test_AUC is {best_test_AUC}')
        print(f'best_test_epoch is {best_test_epoch}')
        print(f'[best_test_AUC, best_test_epoch] is {[best_test_AUC, best_test_epoch]}')
        print(f'test_AUCs_by_epoch_class is {test_AUCs_by_epoch_class}')
        print(f'test_AUCs_by_epoch_class.shape is {test_AUCs_by_epoch_class.shape}')

        combination_str = []
        for comb in combination:
            combination_str.append(str(comb))

        new_row = [comb_index] + list(combination_str) + [best_test_AUC, best_test_epoch] + np.transpose(test_AUCs_by_epoch_class)[best_test_epoch].tolist()
        print(f'new_row is {new_row}')
        sheet1.append(new_row)
        wb.save('/content/drive/My Drive/research/frontiers/performance/performance.xlsx')


if __name__ == "__main__":
    # train 함수를 직접 호출했을 때 실행.
    data_path = "../datasets/severity4_no_space.csv"
    run(data_path)
